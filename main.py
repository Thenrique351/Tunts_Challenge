from request import request_url
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.utils import get_column_letter


dictionary = request_url('https://restcountries.com/v3/all')# Return values from API 
#init xlxs document, reference: https://openpyxl.readthedocs.io/en/stable/tutorial.html
wb = Workbook()
ws = wb.active
thin = Side(border_style="thin", color="D9D9E3")#border color
ws.title = "Countries List"
ws.merge_cells('A1:D1')
#Create title of table
ws['A1'] = 'Countries List'
ws['A1'].font = Font(color="4F4F4F", bold = True, size = 16)#font edit, color, bold and size.
ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws['A1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)#crate border
ws['E1'].border = Border(left = thin)
#Create title of columns

ws['A2'] = 'Name'
ws['A2'].font = Font(color="808080", bold = True, size = 12)
ws['A2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

ws['B2'] = 'Capital'
ws['B2'].font = Font(color="808080", bold = True, size = 12)
ws['B2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

ws['C2'] = 'Area'
ws['C2'].font = Font(color="808080", bold = True, size = 12)
ws['C2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

ws['D2'] = 'Currencies'
ws['D2'].font = Font(color="808080", bold = True, size = 12)
ws['D2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
ws['D2'].alignment = Alignment(vertical="top")

ws.column_dimensions['A'].width = 36.89
ws.column_dimensions['B'].width = 21.96
ws.column_dimensions['C'].width = 9.48
ws.column_dimensions['D'].width = 14.37

# Get datas from dictionary
i = 3 #iterable var
for row in dictionary:
  x = 'A' + str(i) #Cell iterable selection from column A
  ws[x] = row['Name'] #Find information by key name
  ws[x].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  x = 'B' + str(i) #Cell iterable selection from column B
  ws[x] = row['Capital'] #Find information by key name
  ws[x].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  x = 'C' + str(i)
  ws[x] = row['Area']
  ws[x].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  x = 'D' + str(i)
  ws[x] = row['Currencies']
  ws[x].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  i = i + 1
#Save xlsx archive format
wb.save("country_table.xlsx")